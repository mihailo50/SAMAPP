from samapp.models import management_group, model_category, model_manufacturer, asset_model
from django.shortcuts import render, get_object_or_404, HttpResponse, redirect
from samapp.models.asset import Asset
from samapp.models.business_unit import BusinessUnit
import openpyxl
import datetime
from django.contrib import messages
import json


def extract_all_excel(request):
    assets = Asset.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    
    timestamp = datetime.datetime.today().strftime('%Y%m%d[%H%M%S]')
    

    headers = ['Name', 'Model', 'Warranty', 'Business Unit', 'Price', 'Order Number', 'Status']

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    for row_num, asset in enumerate(assets, 2):
        sheet.cell(row=row_num, column=1).value = asset.name
        sheet.cell(row=row_num, column=2).value = str(asset.model)
        sheet.cell(row=row_num, column=3).value = asset.warranty.strftime('%Y-%m-%d') if asset.warranty else None
        sheet.cell(row=row_num, column=4).value = str(asset.business_unit)
        sheet.cell(row=row_num, column=5).value = str(asset.price)
        sheet.cell(row=row_num, column=6).value = asset.order_number
        sheet.cell(row=row_num, column=7).value = asset.get_status_display()

    file_name = f"C:\\SAMWA\\samwa\\samapp\\{timestamp}.xlsx"
    workbook.save(file_name)

    messages.success(request, f"File '{file_name}' extracted successfully.")

    return redirect('all_assets')

def extract_all_json(request):
    assets = Asset.objects.all()

    data = []
    for asset in assets:
        asset_data = {
            'name': asset.name,
            'model': str(asset.model),
            'warranty': asset.warranty.strftime('%Y-%m-%d') if asset.warranty else None,
            'business_unit': str(asset.business_unit),
            'price': str(asset.price),
            'order_number': asset.order_number,
            'status': asset.get_status_display()
        }
        data.append(asset_data)

    timestamp = datetime.datetime.now().strftime("%Y%m%d[%H%M%S]")
    file_path = f"C:\\SAMWA\\samwa\\samapp\\{timestamp}.json"

    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

    messages.success(request, f"File '{timestamp}' extracted successfully.")

    return redirect('all_assets')
        